#!/usr/bin/env python3
"""
Deep dive to decode ICE stats column metrics (metric_1 through metric_6)
by examining the raw Excel files and comparing patterns.
"""

import openpyxl
import json
from pathlib import Path

def extract_column_headers(filepath):
    """Extract column headers from Excel file"""
    print(f"\n{'='*80}")
    print(f"FILE: {filepath.name}")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Get the first sheet
    ws = wb.active
    
    # Look for header rows
    print(f"\nSheet name: {ws.title}")
    print(f"Dimensions: {ws.dimensions}")
    
    # Extract first 10 rows to find headers
    print("\n--- First 10 rows (looking for headers) ---")
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                row_data.append(f"{col_idx}:{cell.value}")
        
        if row_data:
            print(f"Row {row_idx}: {row_data[:10]}")  # First 10 columns
    
    # Look for Florida facilities
    print("\n--- Florida Facilities Sample Data ---")
    for row_idx in range(1, min(50, ws.max_row + 1)):
        facility_name = ws.cell(row=row_idx, column=1).value
        if facility_name and isinstance(facility_name, str):
            if any(x in facility_name.upper() for x in ['FLORIDA', 'HILLSBOROUGH', 'PINELLAS', 'ORANGE', 'POLK', 'SEMINOLE', 'VOLUSIA']):
                print(f"\nRow {row_idx}: {facility_name}")
                # Print all columns
                for col_idx in range(1, min(20, ws.max_column + 1)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        header = ws.cell(row=1, column=col_idx).value or f"Col{col_idx}"
                        print(f"  {header}: {cell.value}")
    
    wb.close()

def analyze_metric_patterns():
    """Analyze patterns across facilities to infer what metrics mean"""
    print("\n" + "="*80)
    print("METRIC PATTERN ANALYSIS")
    print("="*80)
    
    # Load the facilities data
    with open('FY26_detentionStats_02122026_facilities.json', 'r') as f:
        data = json.load(f)
    
    print("\nAnalyzing metric values across different facility types:")
    
    for facility in data['florida_facilities'][:5]:  # First 5 Florida facilities
        name = facility[0]
        facility_type = facility[6] if len(facility) > 6 else 'Unknown'
        metrics = facility[8:14] if len(facility) > 14 else []
        
        print(f"\n{name} ({facility_type}):")
        for i, metric in enumerate(metrics, 1):
            print(f"  Metric {i}: {metric}")

# Main execution
if __name__ == '__main__':
    print("ICE DETENTION STATS COLUMN METRICS DEEP DIVE")
    print("="*80)
    
    # Check all FY Excel files
    excel_dir = Path('./ICE-Detention-Stats')
    if excel_dir.exists():
        for excel_file in sorted(excel_dir.glob('*.xlsx')):
            try:
                extract_column_headers(excel_file)
                break  # Just do first file for now
            except Exception as e:
                print(f"\nError processing {excel_file}: {e}")
    
    # Analyze metric patterns
    try:
        analyze_metric_patterns()
    except Exception as e:
        print(f"\nError analyzing patterns: {e}")
    
    print("\n" + "="*80)
    print("NEXT STEPS:")
    print("1. Compare column positions across multiple FY files")
    print("2. Cross-reference with ICE footnotes (ADP, ALOS, etc.)")
    print("3. Look for official ICE documentation on data schema")
    print("4. Create column mapping document")
    print("="*80)
