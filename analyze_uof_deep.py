#!/usr/bin/env python3
"""Deep dive on UOF data for ICE-relevant patterns."""
import openpyxl
from collections import Counter, defaultdict

wb = openpyxl.load_workbook('./UOF_Request_1-1-2022_through_12-31-2025.xlsx')
ws = wb['New Report']
headers = [str(cell.value) for cell in ws[1]]

rows = []
for row_idx in range(2, ws.max_row + 1):
    row_data = {}
    for col_idx, header in enumerate(headers):
        cell_val = ws.cell(row=row_idx, column=col_idx+1).value
        row_data[header] = str(cell_val) if cell_val is not None else ''
    rows.append(row_data)

corrections_rows = [r for r in rows if 'Correction' in r['Inc: Incident type']]

# === September 2025 spike detail ===
print("=" * 60)
print("SEPTEMBER 2025 CORRECTIONS UOF INCIDENTS (spike month):")
print("=" * 60)
sept_2025 = [r for r in corrections_rows if r['Inc: Occurred date'].startswith('09/') and r['Inc: Occurred date'].endswith('/2025')]
for r in sept_2025:
    eth = r['Cit: Ethnicity'].split('\n')[0]
    age = r['Cit: Age'].split('\n')[0]
    gender = r['Cit: Gender'].split('\n')[0]
    reason = r['UOF: Reason for using force']
    force = r['UOF: Type of force used'].split('\n')[0]
    resist = r['UOF: Citizen resistance'].split('\n')[0]
    inj = r['UOF: Citizen was injured (y/n)']
    div = r['Inc: Org: Division']
    emp = f"{r['Emp: First name'].split(chr(10))[0]} {r['Emp: Last name'].split(chr(10))[0]}"
    rpt = r['Inc: Report # (Use 12 characters AND capital letters)']
    print(f"  {r['Inc: Occurred date']} | {rpt} | {eth} | Age:{age} | {gender} | Reason:{reason} | Force:{force} | Resist:{resist} | Injured:{inj} | Div:{div} | Emp:{emp}")

print(f"\nSept 2025 ethnicity: {Counter(r['Cit: Ethnicity'].split(chr(10))[0] for r in sept_2025)}")

# === Analyze Hispanic UOF in corrections by MONTH for 2024-2025 ===
print("\n" + "=" * 60)
print("HISPANIC CORRECTIONS UOF BY MONTH (2024-2025):")
print("=" * 60)
hisp_monthly = defaultdict(int)
total_monthly = defaultdict(int)
for r in corrections_rows:
    d = r['Inc: Occurred date']
    eth = r['Cit: Ethnicity'].split('\n')[0]
    if d:
        parts = d.split('/')
        if len(parts) == 3:
            m, dd, y = parts
            if y in ('2024', '2025'):
                ym = f"{y}-{m}"
                total_monthly[ym] += 1
                if eth == 'Hispanic':
                    hisp_monthly[ym] += 1

for ym in sorted(total_monthly):
    total = total_monthly[ym]
    hisp = hisp_monthly.get(ym, 0)
    pct = hisp/total*100 if total > 0 else 0
    print(f"  {ym}: {hisp}/{total} Hispanic ({pct:.1f}%)")

# === Look at age=0 entries — what are they? ===
print("\n" + "=" * 60)
print("AGE=0 ENTRIES (potential data issues):")
print("=" * 60)
age_zero = [r for r in rows if r['Cit: Age'].split('\n')[0] == '0']
print(f"Total age=0 records: {len(age_zero)}")
for r in age_zero[:20]:
    print(f"  {r['Inc: Occurred date']} | {r['Inc: Incident type']} | {r['Inc: Org: Division']} | Eth:{r['Cit: Ethnicity'].split(chr(10))[0]} | Gender:{r['Cit: Gender'].split(chr(10))[0]} | Reason:{r['UOF: Reason for using force']}")

# === Check for "Intake" division — could be processing area where ICE transfers happen ===
print("\n" + "=" * 60)
print("INTAKE DIVISION UOF INCIDENTS:")
print("=" * 60)
intake = [r for r in rows if r['Inc: Org: Division'] == 'Intake']
print(f"Total: {len(intake)}")
for r in intake:
    eth = r['Cit: Ethnicity'].split('\n')[0]
    age = r['Cit: Age'].split('\n')[0]
    print(f"  {r['Inc: Occurred date']} | {r['Inc: Report # (Use 12 characters AND capital letters)']} | Eth:{eth} | Age:{age} | Reason:{r['UOF: Reason for using force']} | Force:{r['UOF: Type of force used'].split(chr(10))[0]} | Emp:{r['Emp: First name'].split(chr(10))[0]} {r['Emp: Last name'].split(chr(10))[0]}")

# === Trend: Corrections UOF rate with injury ===
print("\n" + "=" * 60)
print("CORRECTIONS UOF WITH INJURY BY YEAR:")
print("=" * 60)
inj_by_year = defaultdict(int)
total_by_year = defaultdict(int)
for r in corrections_rows:
    d = r['Inc: Occurred date']
    if d:
        year = d.split('/')[-1]
        total_by_year[year] += 1
        if r['UOF: Citizen was injured (y/n)'] == 'Yes':
            inj_by_year[year] += 1
for y in sorted(total_by_year):
    total = total_by_year[y]
    inj = inj_by_year.get(y, 0)
    print(f"  {y}: {inj}/{total} injured ({inj/total*100:.1f}%)")

# === Pre vs Post Trump 2.0 (Jan 20, 2025) corrections UOF ===
print("\n" + "=" * 60)
print("PRE vs POST TRUMP INAUGURATION (Jan 20, 2025) - CORRECTIONS:")
print("=" * 60)
pre_trump = 0
post_trump = 0
pre_trump_hisp = 0
post_trump_hisp = 0
for r in corrections_rows:
    d = r['Inc: Occurred date']
    if d and d.endswith('/2025'):
        parts = d.split('/')
        m, dd = int(parts[0]), int(parts[1])
        eth = r['Cit: Ethnicity'].split('\n')[0]
        if m == 1 and dd < 20:
            pre_trump += 1
            if eth == 'Hispanic':
                pre_trump_hisp += 1
        else:
            post_trump += 1
            if eth == 'Hispanic':
                post_trump_hisp += 1

print(f"  Pre-inauguration (Jan 1-19, 2025): {pre_trump} total, {pre_trump_hisp} Hispanic")
print(f"  Post-inauguration (Jan 20 - Dec 30, 2025): {post_trump} total, {post_trump_hisp} Hispanic")

# === "Prevent Escape" reason — could indicate custody/transfer situations ===
print("\n" + "=" * 60)
print("'PREVENT ESCAPE' CORRECTIONS UOF INCIDENTS:")
print("=" * 60)
escape = [r for r in corrections_rows if r['UOF: Reason for using force'] == 'Prevent Escape']
print(f"Total: {len(escape)}")
for r in escape:
    eth = r['Cit: Ethnicity'].split('\n')[0]
    age = r['Cit: Age'].split('\n')[0]
    print(f"  {r['Inc: Occurred date']} | Eth:{eth} | Age:{age} | Force:{r['UOF: Type of force used'].split(chr(10))[0]} | Div:{r['Inc: Org: Division']}")

# === Analyze Restraint Chair usage — high-severity force ===
print("\n" + "=" * 60)
print("RESTRAINT CHAIR USAGE BY YEAR:")
print("=" * 60)
chair_by_year = defaultdict(int)
for r in rows:
    if 'Restraint Chair' in r['UOF: Type of force used']:
        d = r['Inc: Occurred date']
        if d:
            year = d.split('/')[-1]
            chair_by_year[year] += 1
for y in sorted(chair_by_year):
    print(f"  {y}: {chair_by_year[y]}")

# === Check for "Investigative Detention" reason — could tie to ICE holds ===
print("\n" + "=" * 60)
print("'INVESTIGATIVE DETENTION' CORRECTIONS UOF INCIDENTS:")
print("=" * 60)
inv_det = [r for r in corrections_rows if r['UOF: Reason for using force'] == 'Investigative Detention']
print(f"Total: {len(inv_det)}")
for r in inv_det:
    eth = r['Cit: Ethnicity'].split('\n')[0]
    age = r['Cit: Age'].split('\n')[0]
    print(f"  {r['Inc: Occurred date']} | Eth:{eth} | Age:{age} | Force:{r['UOF: Type of force used'].split(chr(10))[0]} | Div:{r['Inc: Org: Division']}")

# === H2 2025 vs H2 2024 comparison ===
print("\n" + "=" * 60)
print("H2 2024 vs H2 2025 CORRECTIONS COMPARISON:")
print("=" * 60)
h2_2024 = [r for r in corrections_rows if r['Inc: Occurred date'] and r['Inc: Occurred date'].endswith('/2024') and int(r['Inc: Occurred date'].split('/')[0]) >= 7]
h2_2025 = [r for r in corrections_rows if r['Inc: Occurred date'] and r['Inc: Occurred date'].endswith('/2025') and int(r['Inc: Occurred date'].split('/')[0]) >= 7]
print(f"  H2 2024: {len(h2_2024)} corrections UOF")
print(f"  H2 2025: {len(h2_2025)} corrections UOF")
print(f"  Change: {len(h2_2025)-len(h2_2024)} ({(len(h2_2025)-len(h2_2024))/len(h2_2024)*100:.1f}% if base>0)" if len(h2_2024) > 0 else "")
print(f"  H2 2024 Hispanic: {sum(1 for r in h2_2024 if r['Cit: Ethnicity'].split(chr(10))[0] == 'Hispanic')}")
print(f"  H2 2025 Hispanic: {sum(1 for r in h2_2025 if r['Cit: Ethnicity'].split(chr(10))[0] == 'Hispanic')}")

# === Unique incidents (deduplicated by report number) ===
print("\n" + "=" * 60)
print("DEDUPLICATED INCIDENT ANALYSIS (by unique report #):")
print("=" * 60)
# Group by report number
by_report = defaultdict(list)
for r in corrections_rows:
    rpt = r['Inc: Report # (Use 12 characters AND capital letters)']
    by_report[rpt].append(r)

print(f"Unique corrections incident reports: {len(by_report)}")
# Count unique incidents by year
unique_by_year = defaultdict(int)
unique_hisp_by_year = defaultdict(int)
for rpt, recs in by_report.items():
    d = recs[0]['Inc: Occurred date']
    if d:
        year = d.split('/')[-1]
        unique_by_year[year] += 1
        # Check if any record in the incident involves Hispanic citizen
        eths = set()
        for rec in recs:
            for e in rec['Cit: Ethnicity'].split('\n'):
                eths.add(e.strip())
        if 'Hispanic' in eths:
            unique_hisp_by_year[year] += 1

print("\nUnique corrections incidents by year:")
for y in sorted(unique_by_year):
    total = unique_by_year[y]
    hisp = unique_hisp_by_year.get(y, 0)
    pct = hisp/total*100 if total > 0 else 0
    print(f"  {y}: {total} total, {hisp} involving Hispanic citizens ({pct:.1f}%)")

# === Check force escalation patterns ===
print("\n" + "=" * 60)
print("CHEMICAL AGENT / TASER USAGE IN CORRECTIONS BY YEAR:")
print("=" * 60)
chem_by_year = defaultdict(int)
taser_by_year = defaultdict(int)
for r in corrections_rows:
    d = r['Inc: Occurred date']
    if d:
        year = d.split('/')[-1]
        force = r['UOF: Type of force used']
        if 'Chemical Agent' in force:
            chem_by_year[year] += 1
        if 'Taser' in force:
            taser_by_year[year] += 1
print("Chemical Agent:")
for y in sorted(chem_by_year):
    print(f"  {y}: {chem_by_year[y]}")
print("Taser:")
for y in sorted(taser_by_year):
    print(f"  {y}: {taser_by_year[y]}")
