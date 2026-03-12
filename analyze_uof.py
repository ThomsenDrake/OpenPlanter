#!/usr/bin/env python3
"""Comprehensive analysis of Seminole County Sheriff UOF data for ICE investigation relevance."""
import openpyxl
from collections import Counter, defaultdict
import json
import re

wb = openpyxl.load_workbook('./UOF_Request_1-1-2022_through_12-31-2025.xlsx')
ws = wb['New Report']

# Parse all rows
headers = [str(cell.value) for cell in ws[1]]
print("HEADERS:", headers)
print()

rows = []
for row_idx in range(2, ws.max_row + 1):
    row_data = {}
    for col_idx, header in enumerate(headers):
        cell_val = ws.cell(row=row_idx, column=col_idx+1).value
        row_data[header] = str(cell_val) if cell_val is not None else ''
    rows.append(row_data)

print(f"Total rows: {len(rows)}")
print()

# === Basic field value distributions ===
print("=" * 60)
print("INCIDENT TYPES:")
inc_types = Counter(r['Inc: Incident type'] for r in rows)
for k, v in inc_types.most_common():
    print(f"  {k}: {v}")

print()
print("DIVISIONS:")
divisions = Counter(r['Inc: Org: Division'] for r in rows)
for k, v in divisions.most_common():
    print(f"  {k}: {v}")

print()
print("REASON FOR FORCE:")
reasons = Counter(r['UOF: Reason for using force'] for r in rows)
for k, v in reasons.most_common():
    print(f"  {k}: {v}")

print()
print("TYPE OF FORCE USED (first entry per row):")
force_types = Counter(r['UOF: Type of force used'].split('\n')[0] for r in rows)
for k, v in force_types.most_common():
    print(f"  {k}: {v}")

print()
print("CITIZEN RESISTANCE (first entry per row):")
resist = Counter(r['UOF: Citizen resistance'].split('\n')[0] for r in rows)
for k, v in resist.most_common():
    print(f"  {k}: {v}")

print()
print("ETHNICITY (first entry per row):")
eth = Counter(r['Cit: Ethnicity'].split('\n')[0] for r in rows)
for k, v in eth.most_common():
    print(f"  {k}: {v}")

print()
print("GENDER (first entry per row):")
gend = Counter(r['Cit: Gender'].split('\n')[0] for r in rows)
for k, v in gend.most_common():
    print(f"  {k}: {v}")

print()
print("CITIZEN INJURED:")
inj = Counter(r['UOF: Citizen was injured (y/n)'] for r in rows)
for k, v in inj.most_common():
    print(f"  {k}: {v}")

print()
print("CITIZEN CONDITION/INJURY (first entry per row):")
cond = Counter(r['UOF: Citizen condition/injury'].split('\n')[0] for r in rows)
for k, v in cond.most_common(20):
    print(f"  {k}: {v}")

# === Search for ICE/immigration references ===
print()
print("=" * 60)
print("SEARCHING FOR ICE/IMMIGRATION KEYWORDS IN ALL FIELDS:")
ice_keywords = ['ice', 'immigr', 'detainer', 'removal', 'deportat', 'customs', 'enforcement', 'federal', 'ero', '287']
for kw in ice_keywords:
    matches = []
    for i, r in enumerate(rows):
        for h, v in r.items():
            if kw.lower() in v.lower():
                matches.append((i+2, h, v[:100]))
    if matches:
        print(f"\n  Keyword '{kw}' found in {len(matches)} cells:")
        for row_num, header, val in matches[:10]:
            print(f"    Row {row_num}, {header}: {val}")
        if len(matches) > 10:
            print(f"    ... and {len(matches)-10} more")

# === Date analysis ===
print()
print("=" * 60)
print("INCIDENTS BY YEAR-MONTH (Corrections only):")
corrections_by_month = Counter()
all_by_month = Counter()
for r in rows:
    date_str = r['Inc: Occurred date']
    if date_str:
        parts = date_str.split('/')
        if len(parts) == 3:
            month, day, year = parts
            ym = f"{year}-{month}"
            all_by_month[ym] += 1
            if 'Correction' in r['Inc: Incident type']:
                corrections_by_month[ym] += 1

print("\nAll incidents by year:")
year_counts = Counter()
for ym, c in all_by_month.items():
    year = ym.split('-')[0]
    year_counts[year] += c
for y in sorted(year_counts):
    print(f"  {y}: {year_counts[y]}")

print("\nCorrections incidents by year:")
corr_year_counts = Counter()
for ym, c in corrections_by_month.items():
    year = ym.split('-')[0]
    corr_year_counts[year] += c
for y in sorted(corr_year_counts):
    print(f"  {y}: {corr_year_counts[y]}")

# === Corrections vs Patrol breakdown ===
print()
print("=" * 60)
print("CORRECTIONS VS NON-CORRECTIONS BREAKDOWN:")
corrections_rows = [r for r in rows if 'Correction' in r['Inc: Incident type']]
non_corr_rows = [r for r in rows if 'Correction' not in r['Inc: Incident type']]
print(f"  Corrections incidents: {len(corrections_rows)}")
print(f"  Non-corrections incidents: {len(non_corr_rows)}")

print("\nCorrections - Ethnicity:")
eth_corr = Counter(r['Cit: Ethnicity'].split('\n')[0] for r in corrections_rows)
for k, v in eth_corr.most_common():
    print(f"  {k}: {v} ({v/len(corrections_rows)*100:.1f}%)")

print("\nNon-Corrections - Ethnicity:")
eth_nc = Counter(r['Cit: Ethnicity'].split('\n')[0] for r in non_corr_rows)
for k, v in eth_nc.most_common():
    print(f"  {k}: {v} ({v/len(non_corr_rows)*100:.1f}%)")

# === Hispanic breakdown by year ===
print()
print("=" * 60)
print("HISPANIC ETHNICITY UOF INCIDENTS BY YEAR (Corrections):")
hisp_by_year = Counter()
total_by_year = Counter()
for r in corrections_rows:
    date_str = r['Inc: Occurred date']
    eth_val = r['Cit: Ethnicity'].split('\n')[0]
    if date_str:
        year = date_str.split('/')[-1]
        total_by_year[year] += 1
        if 'Hispanic' in eth_val and 'Non-Hispanic' not in eth_val:
            hisp_by_year[year] += 1
for y in sorted(total_by_year):
    total = total_by_year[y]
    hisp = hisp_by_year.get(y, 0)
    print(f"  {y}: {hisp}/{total} ({hisp/total*100:.1f}% Hispanic)")

print("\nHISPANIC ETHNICITY UOF INCIDENTS BY YEAR (All):")
hisp_by_year_all = Counter()
total_by_year_all = Counter()
for r in rows:
    date_str = r['Inc: Occurred date']
    eth_val = r['Cit: Ethnicity'].split('\n')[0]
    if date_str:
        year = date_str.split('/')[-1]
        total_by_year_all[year] += 1
        if 'Hispanic' in eth_val and 'Non-Hispanic' not in eth_val:
            hisp_by_year_all[year] += 1
for y in sorted(total_by_year_all):
    total = total_by_year_all[y]
    hisp = hisp_by_year_all.get(y, 0)
    print(f"  {y}: {hisp}/{total} ({hisp/total*100:.1f}% Hispanic)")

# === Unique incidents (by report #) ===
print()
print("=" * 60)
print("UNIQUE INCIDENT REPORTS:")
unique_reports = set(r['Inc: Report # (Use 12 characters AND capital letters)'] for r in rows)
print(f"  Total unique report numbers: {len(unique_reports)}")
corr_reports = set(r['Inc: Report # (Use 12 characters AND capital letters)'] for r in corrections_rows)
print(f"  Unique corrections reports: {len(corr_reports)}")

# === Employees with most UOF incidents ===
print()
print("=" * 60)
print("TOP 20 EMPLOYEES BY UOF INCIDENT COUNT:")
emp_counts = Counter()
for r in rows:
    emp_id = r['Emp: Employee ID'].split('\n')[0]
    fname = r['Emp: First name'].split('\n')[0]
    lname = r['Emp: Last name'].split('\n')[0]
    emp_key = f"{emp_id} - {fname} {lname}"
    emp_counts[emp_key] += 1
for emp, cnt in emp_counts.most_common(20):
    print(f"  {emp}: {cnt}")

# === Corrections employees specifically ===
print()
print("TOP 20 CORRECTIONS EMPLOYEES BY UOF INCIDENT COUNT:")
emp_corr_counts = Counter()
for r in corrections_rows:
    emp_id = r['Emp: Employee ID'].split('\n')[0]
    fname = r['Emp: First name'].split('\n')[0]
    lname = r['Emp: Last name'].split('\n')[0]
    emp_key = f"{emp_id} - {fname} {lname}"
    emp_corr_counts[emp_key] += 1
for emp, cnt in emp_corr_counts.most_common(20):
    print(f"  {emp}: {cnt}")

# === Monthly trend corrections ===
print()
print("=" * 60)
print("MONTHLY CORRECTIONS UOF TREND:")
for ym in sorted(corrections_by_month):
    print(f"  {ym}: {corrections_by_month[ym]}")

# === Force type analysis for corrections ===
print()
print("=" * 60)
print("FORCE TYPES USED IN CORRECTIONS (all entries, split by newlines):")
corr_force = Counter()
for r in corrections_rows:
    for f in r['UOF: Type of force used'].split('\n'):
        f = f.strip()
        if f:
            corr_force[f] += 1
for k, v in corr_force.most_common():
    print(f"  {k}: {v}")

# === Injury analysis ===
print()
print("=" * 60)
print("INJURY TYPES IN CORRECTIONS:")
corr_inj = Counter()
for r in corrections_rows:
    for f in r['UOF: Citizen condition/injury'].split('\n'):
        f = f.strip()
        if f:
            corr_inj[f] += 1
for k, v in corr_inj.most_common():
    print(f"  {k}: {v}")

# === Check for incidents happening after Trump 2.0 era (Jan 2025+) ===
print()
print("=" * 60)
print("2025 INCIDENTS DETAIL (Corrections - first 30):")
incidents_2025 = [r for r in corrections_rows if r['Inc: Occurred date'].endswith('/2025')]
for i, r in enumerate(incidents_2025[:30]):
    print(f"  {r['Inc: Occurred date']} | {r['Inc: Report # (Use 12 characters AND capital letters)']} | Eth: {r['Cit: Ethnicity'].split(chr(10))[0]} | Reason: {r['UOF: Reason for using force']} | Force: {r['UOF: Type of force used'].split(chr(10))[0]} | Div: {r['Inc: Org: Division']} | Emp: {r['Emp: First name'].split(chr(10))[0]} {r['Emp: Last name'].split(chr(10))[0]}")

print(f"\nTotal 2025 corrections incidents: {len(incidents_2025)}")

# === Look at Q1 2025 specifically (Trump inauguration Jan 20, 2025) ===
print()
print("JANUARY-MARCH 2025 CORRECTIONS UOF:")
for r in corrections_rows:
    d = r['Inc: Occurred date']
    if d.endswith('/2025'):
        parts = d.split('/')
        month = int(parts[0])
        if month <= 3:
            print(f"  {d} | {r['Inc: Report # (Use 12 characters AND capital letters)']} | Eth: {r['Cit: Ethnicity'].split(chr(10))[0]} | Age: {r['Cit: Age'].split(chr(10))[0]} | Reason: {r['UOF: Reason for using force']} | Resist: {r['UOF: Citizen resistance'].split(chr(10))[0]} | Force: {r['UOF: Type of force used'].split(chr(10))[0]}")

# Look at 2025 vs 2024 per-month comparison
print()
print("=" * 60)
print("QUARTERLY UOF COMPARISON (ALL INCIDENTS):")
quarterly = defaultdict(int)
for r in rows:
    d = r['Inc: Occurred date']
    if d:
        parts = d.split('/')
        if len(parts) == 3:
            m, dd, y = parts
            q = (int(m)-1)//3 + 1
            quarterly[f"{y}-Q{q}"] += 1
for k in sorted(quarterly):
    print(f"  {k}: {quarterly[k]}")

# === CRITICAL: Check for any 287(g) or immigration hold related divisions ===
print()
print("=" * 60)
print("ALL UNIQUE DIVISIONS:")
all_divs = set(r['Inc: Org: Division'] for r in rows)
for d in sorted(all_divs):
    print(f"  '{d}'")

print()
print("ALL UNIQUE INCIDENT TYPES:")
all_types = set(r['Inc: Incident type'] for r in rows)
for t in sorted(all_types):
    print(f"  '{t}'")

# Check date range
print()
print("=" * 60)
dates_parsed = []
for r in rows:
    d = r['Inc: Occurred date']
    if d:
        parts = d.split('/')
        if len(parts) == 3:
            dates_parsed.append(d)
dates_sorted = sorted(dates_parsed, key=lambda x: (x.split('/')[2], x.split('/')[0], x.split('/')[1]))
print(f"Date range: {dates_sorted[0]} to {dates_sorted[-1]}")
print(f"Total records: {len(rows)}")
