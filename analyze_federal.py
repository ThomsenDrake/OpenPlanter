import openpyxl, re
from collections import Counter, defaultdict
wb = openpyxl.load_workbook('PRR_25-6990_Clearview_AI.xlsx', data_only=True)
ws = wb['PRR Clearview Searches']

# All federal agency references in titles
print('=== ALL FEDERAL AGENCY REFERENCES ===')
fed_keywords = {
    'HSI': ['hsi'],
    'USPIS': ['uspis'],
    'MBI': ['mbi ','mbi/','mbi-'],
    'FBI': ['fbi'],
    'ATF': ['atf'],
    'DEA': ['dea ','dea-'],
    'USMS': ['usms','marshal'],
    'Federal': ['federal'],
    'CFIX': ['cfix'],
    'HSIN': ['hsin'],
    'Secret Service': ['secret service','usss'],
    'DHS': ['dhs ','dhs-'],
}

fed_records = defaultdict(list)
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    title = str(row[1]) if row[1] else ''
    all_text = ' '.join([str(v) if v else '' for v in row]).lower()
    for agency, kws in fed_keywords.items():
        for kw in kws:
            if kw in all_text:
                fed_records[agency].append((row_idx, title[:70], str(row[4])[:10] if row[4] else ''))
                break

for agency in sorted(fed_records.keys()):
    records = fed_records[agency]
    print(f'\n{agency} ({len(records)} records):')
    for r in records:
        print(f'  Row {r[0]}: {r[1]} | {r[2]}')

# Undercover personnel exemption records
print('\n\n=== UNDERCOVER PERSONNEL EXEMPTION (F.S. 119.071(4)(c)) - ALL RECORDS ===')
uc_count = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    creator = str(row[2]) if row[2] else ''
    if '119.071(4)(c)' in creator:
        uc_count += 1
        print(f'Row {row_idx}: Title={row[1]} | Created={str(row[4])[:10]} | Comments={row[9]}')
print(f'\nTotal undercover exemption records: {uc_count}')

# Monthly volume over time
print('\n=== MONTHLY VOLUMES ===')
monthly = defaultdict(int)
for row in ws.iter_rows(min_row=2, values_only=True):
    created = row[4]
    if created:
        month_key = created.strftime('%Y-%m')
        monthly[month_key] += 1

for month in sorted(monthly.keys()):
    print(f'  {month}: {monthly[month]}')
